#Purpose: To efficiently produce a balanced pooling structure for freemuxlet-based scRNAseq experiments.
#These experiments require that only one sample from a given participant is present in a given pool, and that all of the samples from a participant appear in a unique combination of all available pools.
#This program takes as input a .xlsx file with the first column identifying the participant, and the second column identifying the sample.
#The column must have a header in row one but the exact text of the header is not important.
#Any .xlsx file that is in the same directory as this program can be selected as the source file.
#The output is a .xlsx file that details the pooling strategy.
#The value of this program is that it will balance features of each participant and or each sample so they are distributed evenly across the pools.
#If additional features are to be considered they should be included as columns beyond the two that are required for the program to function.
#Sample specific features must have column heads that contain the word 'sample' (not case specific).
#Participant specific features must have column heads that contain the word 'participant' (not case specific).
#This balancing is accomplished by scoring distinct combinations and permutations of possible pool assignments for each participant and associated sample respectively.
#The score is increase by the presence of other samples in the pool that have the same attributes of those being considerd for addition to the pool.
#After all possible combinations and permutations have been checked. The combo or perm with the lowest score is selected.
import pandas as pd
import time
import os
from itertools import combinations, permutations
from datetime import datetime
from openpyxl import load_workbook


#Sets the working directory to the folder in which this .py file resides.
def directory_set_and_folder_creation():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    #Creates destination folders inside this working directory if they do not exist.
    if not os.path.exists(os.path.dirname(os.path.abspath(__file__)) + '/Output'):
        os.makedirs(os.path.dirname(os.path.abspath(__file__)) + '/Output')

#Docker directory access (must be defined in the run command)
# def directory_set_and_folder_creation():
#     os.chdir('/mnt/mydata')
#     if not os.path.exists('/mnt/mydata/Output'):
#         os.makedirs('/mnt/mydata/Output')

#Allows the user to identify the file that has sample information for pooling.
def file_identification():
    #Initializes a dictionary that is displayed to the user so that they can select the appropriate file.
    file_name_dict = {}
    file_index = 1
    #Iterates through the working directory and identifies any .xlsx files as candidates for selection.
    for file in os.listdir(os.getcwd()):
        if '.xlsx' in file:
            print(f'{file_index}: {file}')
            file_name_dict[file_index] = file
            file_index += 1
    #Prompts the user to enter the integer value assoicated with the desired file.
    #If the entry is not a valid integer option the user is prompted again.
    #The loop is broken by a valid entry.
    while True:
        selected_file = input("Enter the integer associated with the sample list file.\n")
        try:
            selected_file = int(selected_file)
            if selected_file <= len(file_name_dict):
                break
            else:
                print('This value is not an option given the provided list!\n')
        except:
            print('The value entered is not recognized as an integer!\n')
    sample_list_file = file_name_dict[selected_file]
    return sample_list_file

#This function generates the dictionary that is used throughout the rest of the program.
#This dictionary becomes heavily nested so it is critical to understand its structure if this code is to be modified. 
def pad_generator(s_list):
    #Initializes the central dictionary for the entire program.
    pool_assignment_dictionary = {}
    #Brings in all of the data from the identified source file.
    df = pd.read_excel(s_list)
    #Generates a list of all of the column heads in the data frame that is used for critical data allocation.
    column_list = df.columns.tolist()
    #Generates a list of all of the unique participants in the source file. Per the earlier direction above, the participants must be listed in the first column.
    unique_participants = df[column_list[0]].unique().tolist()
    #Iterates through the list of unique participants and creates an entry in the pool_assignment_dictionary for each.
    #The value for each of these participant-keys is a dictionary with two keys that also have dictionaries as values.
    for participant in unique_participants:
        pool_assignment_dictionary[participant] = {'Sample_IDs': {}, "Participant_Attributes": {}}
        #Initialize a counter that tracks the number of samples for each participant individually.
        #This value is assigned to the Sample_Count key at the before the function loops to the next participant and value is reset.
        participant_sample_count = 0
        #This loop checks each sample in the input file to see if it contains a sample that belongs the the participant being checked.
        #If the sample is associated with the participant, then it is added to the dictionary under the Sample_IDs key.
        #The value for the sample-key is the final dictionary in its tree which contains a key for the samples future pool assignment (Pool), and will also contain any sample attributes.
        for index, row in df.iterrows():
            if row[column_list[0]] == participant:
                pool_assignment_dictionary[participant]['Sample_IDs'][row[column_list[1]]] = {'Pool': 0}
                participant_sample_count += 1
                #If additional sample or participant-specific criteria were included in the provided sample data spreadsheet those additional columns will be considered here.
                if len(column_list) > 2:
                    for attribute in column_list[2:]:
                        #As described earlier additional columns beyond the participant and timepoint columns are considered in their relation to specific participants or specific samples here.
                        if 'sample' in attribute.lower():
                            #Sample specific attributes are added into the same dictionary as the future pool assignemnt for that sample.
                            pool_assignment_dictionary[participant]['Sample_IDs'][row[column_list[1]]][attribute] = [row[attribute]][0]
                        elif 'participant' in attribute.lower():
                            #Participant level attributes are added into a nested Participant_Attributes dictionary that exists at the same level as the Sample_IDs dictionary for the participant.
                            pool_assignment_dictionary[participant]["Participant_Attributes"][attribute] = [row[attribute]][0]
        #Finally the sample count is added to the participant's dictionary.
        pool_assignment_dictionary[participant]['Sample_Count'] = participant_sample_count
    return pool_assignment_dictionary

#This function determines the minimum number of pools to include all the samples listed and have one healthy control that spans all of the pools.
#It uses the sample count from each participant to determine the participant with the greatest number of samples and adds one on top of that so the HC is the only sample in ALL pools.
def minimum_pools_with_control_in_all_pools(min_pools_pad):
    min_pools = 0
    sample_count_per_participant_dict = {}
    for key in min_pools_pad:
        if min_pools_pad[key]['Sample_Count'] in sample_count_per_participant_dict:
            sample_count_per_participant_dict[min_pools_pad[key]['Sample_Count']] += 1
        else:
            sample_count_per_participant_dict[min_pools_pad[key]['Sample_Count']] = 1
        if min_pools_pad[key]['Sample_Count'] > min_pools:
            min_pools = min_pools_pad[key]['Sample_Count']
    min_pools += 1
    #This loop is an essential check to see if there are too many participants with a given number of samples such that
        #together they exceed the combinatoric capacity of the current number of alotted pools.
    for participant_sample_count in sample_count_per_participant_dict:
        #This while loop is needed to check if more than a single incrementation in pool count is needed to accomodate the number of
            #individual participants with a given sample count.
        while True:
            if sample_count_per_participant_dict[participant_sample_count] > len(list(combinations(range(min_pools),participant_sample_count))):
                min_pools += 1
            else:
                break
    return min_pools

#This function is a simple user input point that allows the user to exceed the minimum number of pools if desired.
def user_determined_pool_count():
    while True:
        pool_count = input('\n~~~~~~~~~~~~~~~\nIf you have a desired number of pools\nenter the interger now.\nElse press <return>.\n')
        if pool_count != '':
            try:
                pool_count = int(pool_count)
                return pool_count
            except:
                print(f'{pool_count} is not recognized as an integer!')
        return False

#This function compares the user input with the minimum number of required pools and will override the user input if needed to accomodate all of the samples.
def pool_count_determination(i_c_g_pad):
    pools = user_determined_pool_count()
    if not pools:
        pools = minimum_pools_with_control_in_all_pools(i_c_g_pad)
    elif pools < minimum_pools_with_control_in_all_pools(i_c_g_pad):
        print('There are too many samples for one or more patients to use that few sample pools!')
        pools = minimum_pools_with_control_in_all_pools(i_c_g_pad)
    return pools

#This function consists of a complex scoring system for pool size and participant level attributes.
#This system is used to determine the ideal combination of pools for each participant such that the pool sizes and participant level attributes are as balanced as possible.
def combo_selection(c_s_pad, number_pools):
    #Initialize dictionaries that help to balance pool size and participant-level attributes respectively.
    pool_counter_dict = {}
    participant_attribute_counter = {}
    #Initialize a list that makes sure no two participants are assigned to the exact same combination of pools.
    unique_combo_list = []
    #This loop begins to populate the two dictionaries created above with entries for each of the pools in the experiment.
    for i in range(1, number_pools + 1):
        #The pool counter is a simple value that tracks the size of a given pool.
        pool_counter_dict[i] = 0
        #The participant attribute counter can track as many participant attributes as are provided in the sample data spreadsheet through a nested dictionary.
        participant_attribute_counter[i] = {}
        #This loop populates the participant attribute counters nested dictionary with keys for each new attribute value of each attribute type present for all of the participants.
        #All of the count values are set to zero.
        for participant in c_s_pad:
            for p_attribute in c_s_pad[participant]["Participant_Attributes"]:
                if not p_attribute in participant_attribute_counter[i]:
                    participant_attribute_counter[i][p_attribute] = {}
                #attribute counter dict
                                             #pool
                                                #attribute
                                                             #This region identifies the attribute value##################
                                                             #The core dict
                                                                     #participant
                                                                                  #Attribute dictionary
                                                                                                            #attribute
                participant_attribute_counter[i][p_attribute][c_s_pad[participant]["Participant_Attributes"][p_attribute]] = 0
    #At this stage the process of selecting pool combinations for each of the participants begins.
    for participant in c_s_pad:
        #An empty tuple is created to hold the best combination at any point in the iteration process.
        best_combo = ()
        #The low score variable is set to an unreasonably high value for the first comparison.
        lowest_combo_score = 1000000
        #The itertools combinations module is used to generate all combos for the sample count from each participant.
        #This list is iterated through to check all of the combinations produced for the best possible score.
        for combo in combinations(range(1, number_pools + 1), c_s_pad[participant]['Sample_Count']):
            combo_score = 0
            #If the combination has not already been used then the combo is scored and checked against the best_combo.
            if not combo in unique_combo_list:
                #The first part of the score is the total number of samples that are already present in the pools identified by the combo.
                #
                ###
                #####If one wanted to weight pool balance more substantially, a multiplier could be added to the line "combo_score += pool_counter_dict[pool_integer]"
                ###
                #
                for pool_integer in combo:
                    combo_score += pool_counter_dict[pool_integer]
                    #The next part of the score is based on the number of samples in each pool that match participant specific criteria.
                    #These could be collectively weighted (but that doesn't make sense if you are also weighting pool counts),
                        #or they could be ###Iteratively Weighted### to allow prioritization of attributes present in earlier columns over those present in later columns.
                    for p_att_key in participant_attribute_counter[pool_integer]:
                        combo_score += participant_attribute_counter[pool_integer][p_att_key][c_s_pad[participant]["Participant_Attributes"][p_att_key]]
                #A simple comparison will reassign best combo to the current combo if the score is lower.
                #lowest combo score is also set to the lower value if true.
                if combo_score < lowest_combo_score:
                    lowest_combo_score = combo_score
                    best_combo = combo
            else:
                continue
        #The combo that is selected after the list is fully iterated is added to the unique combo list to prevent duplicate entry.
        unique_combo_list.append(best_combo)
        #The values for the counter dictionaries are then updated to serve as a reference for subsequent participant samples.
        for pool_integer in best_combo:
            pool_counter_dict[pool_integer] += 1
            #Updating the value for the participant level attibutes is... complicated because of the nested dictionaries involved in both the reference and destination for the data update.
            for p_att_key in participant_attribute_counter[pool_integer]:
                #counter dict
                                             #pool being updated
                                                           #attribute being updated
                                                                      #This entire region identifies the attribute value########
                                                                      #It will provide the value of the attribute ##############
                                                                      #the core dictionary
                                                                               #participant
                                                                                           #The attributes for that participant
                                                                                                                     #the attribute in question
                participant_attribute_counter[pool_integer][p_att_key][c_s_pad[participant]["Participant_Attributes"][p_att_key]] += 1
        #The chosen combination is added to the participants entry in the core dictionary of this program.
        c_s_pad[participant]["Selected_Combo"] = best_combo
    return(c_s_pad, pool_counter_dict)

#This function contains a scoring system for all of the permutations of the combination selected above for each participant.
#It will test these permutations and assign one that best balances sample-specific attributes across the pools.
def permutation_selection(ps_pad, perm_pool_count):
    #Initialize a dictionary that is used to track the counts for sample specific attributes in each pool.
    sample_attribut_counter = {}
    for i in range(1, perm_pool_count + 1):
        sample_attribut_counter[i] = {}
        #Iterates through every attribute for every sample for every participant, disregards the pool assignment, and adds any new attribute value to the dictionary with a set value of 0.
        for participant in ps_pad:
            for sample in ps_pad[participant]['Sample_IDs']:
                for s_attribute in ps_pad[participant]['Sample_IDs'][sample]:
                    if s_attribute == 'Pool':
                        continue
                    else:
                        if not s_attribute in sample_attribut_counter[i]:
                            sample_attribut_counter[i][s_attribute] = {}
                        #counter dict
                                               #Pool
                                                  #attribute
                                                               #This region identifies the attribute value############
                                                               #The core dictionary
                                                                      #participant
                                                                                   #Sample ID dict
                                                                                                 #specific sample
                                                                                                         #attribute
                        sample_attribut_counter[i][s_attribute][ps_pad[participant]['Sample_IDs'][sample][s_attribute]] = 0
    for participant in ps_pad:
        #initializes an empty tuple in which to store the best permutation during iteration.
        best_perm = ()
        #Sets an unreasonably high value for the initial comparison.
        lowest_perm_score = 1000000
        #Iterates through a list of permutations created based on the best combination identified by the previous function.
        for perm in permutations(ps_pad[participant]['Selected_Combo']):
            #Sets the score of each permutation to zero prior to comparison.
            perm_score = 0
            #Uses the zip function to join the permutation and list of sample IDs for joint iteration.
            #These two iterables must be the same length based on how the program is designed.
            for pool_integer, sample in zip(perm, ps_pad[participant]['Sample_IDs']):
                #Iterates through each sample attribute in the counter dictionary.
                for s_att_key in sample_attribut_counter[pool_integer]:
                    #Adds the number of samples that exist in the relevant pool with matching attribute values to the sample being considered for the pool to the permutation score.
                                  #counter dict
                                                         #pool
                                                                       #attribute
                                                                                  #This region identifies the attribute value###########
                                                                                  #core dict
                                                                                         #participant
                                                                                                      #Sample Dict
                                                                                                                    #specific sample
                                                                                                                            #attribute
                    perm_score += sample_attribut_counter[pool_integer][s_att_key][ps_pad[participant]['Sample_IDs'][sample][s_att_key]]
            #A simple comparison to determine if the current permuation and its score should replace the existing values
            if perm_score < lowest_perm_score:
                lowest_perm_score = perm_score
                best_perm = perm
        #Iteratively updates the attribute counter dict with new counts for each attribute in each pool.
        #Value increment is structured in the same fashion as the value check in the previous loop.
        for pool_integer, sample in zip(best_perm, ps_pad[participant]['Sample_IDs']):
            for s_att_key in sample_attribut_counter[pool_integer]:
                sample_attribut_counter[pool_integer][s_att_key][ps_pad[participant]['Sample_IDs'][sample][s_att_key]] += 1
        #The selected permutation is then noted in the core dictionary.
        ps_pad[participant]['Selected_Perm'] = best_perm
    return(ps_pad)

#This function uses the values of a participants assigned permuation to update the pool assignment for each of the participants samples.
def sample_assignment_to_pools(sap_pad):
    for participant in sap_pad:
        for pool, sample in zip(sap_pad[participant]['Selected_Perm'], sap_pad[participant]['Sample_IDs']):
            sap_pad[participant]['Sample_IDs'][sample]['Pool'] = pool
    return sap_pad

#This function generates a data frame using pandas and exports it to an excel file in the output folder that exists in the same directory as this function.
def output_pooling_table(output_pad, total_number_pools):
    #Identifies the current date and time for use in naming the file in a unique manner.
    today = datetime.now()
    #Initializes a list that will be used to create columns for the data frame.
    #Pool will be the first column.
    column_head_list = ['Pool']
    #The participants will be added as subsequent columns.
    for participant in output_pad:
        column_head_list.append(participant)
    #The healthy control sample is the last column by convention.
    column_head_list.append('HC')
    #Create the pandas dataframe with the defined columns and enough rows for all of the pools.
    df = pd.DataFrame(columns = column_head_list, index=range(total_number_pools))
    #Initialize an index variable to target specific coordinates for update.
    index = 0
    #Loop through the Pool and HC columns to update each with the appropriate values.
    for i in range(1, total_number_pools + 1):
        df.loc[index, 'Pool'] = i
        df.loc[index, 'HC'] = 'HC'
        index += 1
    #Set core data entry point values and definitions beneath the pooling strategy.
    df.loc[(total_number_pools + 1), 'Pool'] = " "
    df.loc[(total_number_pools + 2), 'Pool'] = 'Study:'
    df.loc[(total_number_pools + 3), 'Pool'] = " "
    df.loc[(total_number_pools + 4), 'Pool'] = 'Healthy Control:'
    df.loc[(total_number_pools + 5), 'Pool'] = " "
    df.loc[(total_number_pools + 6), 'Pool'] = '10X Version:'
    df.loc[(total_number_pools + 7), 'Pool'] = " "
    df.loc[(total_number_pools + 8), 'Pool'] = 'Cells Loaded:'
    df.loc[(total_number_pools + 9), 'Pool'] = " "
    df.loc[(total_number_pools + 10), 'Pool'] = 'ACK Lysed Samples Highlighted Red'
    df.loc[(total_number_pools + 11), 'Pool'] = " "
    df.loc[(total_number_pools + 12), 'Pool'] = 'Potentially Lost Samples Highlighted Grey'

    #Loop through the core dictionary one more time to place the sample names in their participants column and assigned pools row.
    for participant in output_pad:
        for sample, sample_info_dict in output_pad[participant]['Sample_IDs'].items():
            df.loc[sample_info_dict['Pool'] - 1, participant] = sample
    #Set a name for the file based on the current date and time.
    global file_name 
    file_name = f'Output/Experiment_Strategy_{today.strftime("Date_%Y_%m_%d_Time_%H_%M_%S").replace("/","_")}.xlsx'
    #Export the data frame to an excel file for use in further experimental planning.
    df.to_excel(file_name, index=False, sheet_name='Pooling Strategy')


#Adds a new tab to the excel file that will give users a place to add in sample counts, and will output key experiment data.
def output_counting_table(output_2_pad, tot_pool_count, num_samples_per_pool):
    #Initialize a list with the HC sample accounted for in the first position and assigned to pool 0 for sorting purposes.
    output_list = [['HC', 'HC', 0]]
    for participant in output_2_pad:
        #Create a list for each participant sample [0] will hold 'Sequence ID', [1] is 'Sample ID', and [2] is 'Pool'
        for sample in output_2_pad[participant]['Sample_IDs']:
            participant_list = [0, sample, output_2_pad[participant]['Sample_IDs'][sample]['Pool']]
            #Add each individual sample list to the larger output list.
            output_list.append(participant_list)
    #Sort the list by pool.
    output_list.sort(key = lambda x: x[2])
    #Assign a sequence ID to each sample in the sorted output list at index [0] of the sample list.
    i = 1
    for item in output_list:
        if item[0] == 'HC':
            continue
        item[0] = i
        i += 1
    
    #Initialize a data frame with all of the columns needed for counting samples, proceeding with the multiplex, and tracking remaining cell mass for refreezing.
    df = pd.DataFrame(columns=['Sequence ID', #A
                               'Sample ID', #B
                               'Pool', #C
                               'Sequence ID.', #D
                               'Viability (%)', #E
                               'Total cells (e6/mL)', #F
                               'Viable cells (e6/mL)', #G
                               'Number of Cells Transferred (e6/mL)', #H
                               'Volume to Tx (uL)', #I
                               'Sequence ID..', #J
                               'Pool', #K
                               'Study-Pt-Tp', #L
                               'Remaining Cell Mass + e6 *Assumes 500uL Suspension Volume', #M
                               'Refrozen PBMC', #N
                               'Sequence ID...', #O
                               'Today\'s Date',]) #P
    
    #At each index add the appropriate data to each column.
    #Sequence ID is repeated throughout the data frame with the addition of '.' characters for column targeting.
    #However the subsequent Sequence ID columns all reference the first value in case of any user alterations.
    #i+1 is seen throughout because data frame row indicies and excel row indicies are offset by 1
    i = 1
    for item in output_list:
        df.loc[i, 'Sequence ID'] = item[0]
        df.loc[i, 'Sample ID'] = item[1]
        df.loc[i, 'Pool'] = item[2]
        df.loc[i, 'Sequence ID.'] = f'=A{i+1}'
        #Multiplies the total cell count by the percentage viable and divides by 100.
        df.loc[i, 'Viable cells (e6/mL)'] = f'=IF(F{i+1}<>"",F{i+1}*E{i+1}/100,"")'
        #Uses the volume transferred and the cell concentration to track how many cells from each sample are pooled.
        df.loc[i, 'Number of Cells Transferred (e6/mL)'] = f'=IF(F{i+1}<>"",I{i+1}/1000*F{i+1},"")'
        #Takes the number of samples in this samples pool and divides 1.1 by that number to determine
            #the e6 cell count that should be added.
        #References the tracking dictionary created when generating ideal combinations and adds 1 due to the HC sample.
        if item[2] == 0:
            df.loc[i, 'Volume to Tx (uL)'] = f'=IF(F{i+1}<>"",1000*((0.55/{((len(output_list)-1)/tot_pool_count) + 1})/F{i+1}),"")'
        else:
            df.loc[i, 'Volume to Tx (uL)'] = f'=IF(F{i+1}<>"",1000*((1.1/{num_samples_per_pool[item[2]] + 1})/F{i+1}),"")'
        df.loc[i, 'Sequence ID..'] = f'=A{i+1}'
        df.loc[i, 'Pool'] = item[2]
        #The newer '=CONCAT()' function leads to a prepended '@' character and disrupts function therefore '=CONCATENATE()' is used.
        #"\'Pooling Strategy\'!B{tot_pool_count + 3}" references the exact cell in which the study ID is to be recorded.
        df.loc[i, 'Study-Pt-Tp'] = f'=IF(\'Pooling Strategy\'!B{tot_pool_count + 3}<>"",concatenate(\'Pooling Strategy\'!B{tot_pool_count + 3},"-{item[1]}"),"")'
        #Calculates and rounds the cell mass that remains after the needed cell suspension is used in the experiment.
        #This number will be used to help create the refrozen sample labels.
        df.loc[i, 'Remaining Cell Mass + e6 *Assumes 500uL Suspension Volume'] = f'=IF((F{i+1})>1,concatenate(ROUND((F{i+1}/2)-0.5,1),"e6"),"")'
        df.loc[i, 'Refrozen PBMC'] = 'Refrozen PBMC'
        df.loc[i, 'Sequence ID...'] = f'=A{i+1}'
        df.loc[i, 'Today\'s Date'] = '=today()'
        i += 1
    #Special existing file access function.
    book = load_workbook(file_name)
    #Mode 'a' will append instead of overwrite ('w')
    writer = pd.ExcelWriter(file_name, engine = 'openpyxl', mode = 'a')
    writer.Workbook = book
    df.to_excel(writer, index=False, sheet_name='Sample Counts')
    writer.close()

#Adds a normalization tab to the existing excel file that is used at the end of the experiment to prepare samples for the single cell sequencing process.
def output_normalization_table(tot_pools):
    df = pd.DataFrame(columns=['Pool',
                               'Total Cells (e6/mL)',
                               'Viability (%)',
                               'Initial Volume (uL)',
                               'Final Volume (uL)',
                               'Media to Add to Make V2 (uL)',
                               'Target Cell Conc (e6/mL)',
                               'Volume to Load (uL)',
                               'Cells to Load (e6)',
                               'cDNA Concentration (ng/uL)'])
    #Adds an entry for each pool.
    for i in range(1, tot_pools + 1):
        df.loc[i, 'Pool'] = i
        #Calculates the final cell suspension volume that will contain the desired number of cells in the volume called for in the 10X protocol.
        df.loc[i, 'Final Volume (uL)'] = f'=IF(G{i+1}<>"",(B{i+1}*D{i+1})/G{i+1},"")'
        #Determines the volume that should be added to reach the final volume from the existing/initial cell suspension volume.
        df.loc[i, 'Media to Add to Make V2 (uL)'] = f'=IF(E{i+1}<>"",E{i+1}-D{i+1},"")'
        #Based on the loading volume and desired cell mass.
        df.loc[i, 'Target Cell Conc (e6/mL)'] = f'=IF(AND(ISNUMBER(H{i+1}),ISNUMBER(I{i+1})),I{i+1}/H{i+1}*1000,"")'
        df.loc[i, 'Volume to Load (uL)'] = 'Enter 10X Version-Specific Cell Suspension Loading Volume'
        df.loc[i, 'Cells to Load (e6)'] = 'Enter Desired Cell Loading'
        i += 1
    #Same as annotated above.
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine = 'openpyxl', mode = 'a')
    writer.Workbook = book
    df.to_excel(writer, index=False, sheet_name='Normalization')
    writer.close()

def main():
    start_time = time.time()
    directory_set_and_folder_creation()
    user_sample_list = file_identification()
    pad = pad_generator(user_sample_list)
    pool_count = pool_count_determination(pad)
    pad_selected_combo, samples_per_pool_dict = combo_selection(pad, pool_count)
    permutation_selection(pad_selected_combo, pool_count)
    pad_pools_assigned = sample_assignment_to_pools(pad_selected_combo)
    output_pooling_table(pad_pools_assigned, pool_count)
    output_counting_table(pad_pools_assigned, pool_count, samples_per_pool_dict)
    output_normalization_table(pool_count)
    print(f"\n--- {(time.time() - start_time):.2f} seconds ---\n")
main()